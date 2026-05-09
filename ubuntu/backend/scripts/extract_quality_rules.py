from __future__ import annotations

import argparse
import json
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import]


SHEET_GROUPS: dict[str, list[str]] = {
    "administrativo": ["1. Administrativo (2)", "1. Administrativo"],
    "juridico": ["2. Juridico", "2. Jur?dico"],
    "fisico": ["3. Fisico"],
    "economico": ["4. Economico"],
    "topologica": ["5. Topologica"],
    "geografico": ["6. Geografico", "6. Geogr?fico"],
    "topografico": ["7. Topografico"],
    "novedades": ["8. Novedades"],
}

FIELD_ALIASES: dict[str, list[str]] = {
    "id": ["id", "item"],
    "description": ["descripcion", "descripci?n"],
    "classes": ["clases_asociadas", "clases"],
    "variables": ["variable_asociada", "variables_asociadas"],
    "technical_rule": ["regla_lenguaje_tecnico_ilc", "regla_lenguaje_tecnico"],
    "component": ["componente"],
}

KNOWN_HEADER_NAMES = {alias for aliases in FIELD_ALIASES.values() for alias in aliases}


def normalize_label(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    normalized = unicodedata.normalize("NFD", text)
    normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    normalized = normalized.replace(" ", "_")
    return normalized


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def resolve_columns(header_row: list[str | None]) -> dict[str, int]:
    normalized_headers = [normalize_label(cell) for cell in header_row]
    positions: dict[str, int] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized_headers:
                idx = normalized_headers.index(alias)
                positions[field] = idx
                break
    return positions


def detect_header(ws, scan_rows: int = 6) -> tuple[list[str | None], int]:
    scanned = list(ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True))
    if not scanned:
        return [], 0
    width = max(len(row) for row in scanned)
    header: list[str | None] = []
    header_row_index = 0
    for col in range(width):
        chosen_value = None
        chosen_idx = 0
        fallback_value = None
        fallback_idx = 0
        for idx, row in enumerate(scanned, start=1):
            if col >= len(row):
                continue
            value = row[col]
            if not isinstance(value, str):
                continue
            if not value.strip():
                continue
            normalized = normalize_label(value)
            if normalized and normalized in KNOWN_HEADER_NAMES:
                chosen_value = value
                chosen_idx = idx
                break
            if fallback_value is None:
                fallback_value = value
                fallback_idx = idx
        if chosen_value is None and fallback_value is not None:
            chosen_value = fallback_value
            chosen_idx = fallback_idx
        header.append(chosen_value)
        header_row_index = max(header_row_index, chosen_idx)
    if header_row_index == 0:
        header_row_index = 1
    return header, header_row_index


def extract_entries(ws, slug: str) -> list[dict[str, Any]]:
    header_row, header_row_index = detect_header(ws)
    columns = resolve_columns(header_row)
    data_start = header_row_index + 1

    entries: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        def pick(field: str, fallback_idx: int | None) -> Any:
            idx = columns.get(field, fallback_idx if fallback_idx is not None else -1)
            if idx is None or idx < 0:
                return None
            if idx >= len(row):
                return None
            return row[idx]

        rule_id = normalize_text(pick("id", 0))
        description = normalize_text(pick("description", 1))
        if not rule_id and not description:
            continue
        entry = {
            "id": rule_id,
            "description": description,
            "classes": normalize_text(pick("classes", 2)),
            "variables": normalize_text(pick("variables", 3)),
            "technical_rule": normalize_text(pick("technical_rule", 4)),
            "component": normalize_text(pick("component", 5)),
            "sheet_slug": slug,
        }
        entries.append(entry)
    return entries


def gather_rules(source: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(source, data_only=True)
    grouped: dict[str, list[dict[str, Any]]] = {}
    normalized_lookup = {
        normalize_label(ws.title): ws for ws in workbook.worksheets if normalize_label(ws.title)
    }
    for slug, names in SHEET_GROUPS.items():
        worksheet = None
        for candidate in names:
            key = normalize_label(candidate)
            if key and key in normalized_lookup:
                worksheet = normalized_lookup[key]
                break
        if not worksheet:
            continue
        entries = extract_entries(worksheet, slug)
        if entries:
            grouped[slug] = entries
    return grouped


def write_outputs(target_dir: Path, grouped: dict[str, list[dict[str, Any]]]) -> list[Path]:
    target_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    combined: list[dict[str, Any]] = []
    for slug, entries in grouped.items():
        entries_sorted = sorted(entries, key=lambda item: (item.get("id") or "", item.get("description") or ""))
        payload = {
            "slug": slug,
            "rule_count": len(entries_sorted),
            "rules": entries_sorted,
        }
        destination = target_dir / f"{slug}.json"
        destination.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        written.append(destination)
        combined.extend(entries_sorted)

    combined_path = target_dir / "all_rules.json"
    combined_payload = {
        "total_rules": len(combined),
        "groups": sorted(
            (
                {
                    "slug": slug,
                    "rule_count": len(entries),
                }
                for slug, entries in grouped.items()
            ),
            key=lambda item: item["slug"],
        ),
        "rules": combined,
    }
    combined_path.write_text(json.dumps(combined_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    written.append(combined_path)
    return written


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extrae reglas de calidad desde un Excel oficial.")
    parser.add_argument("--source", type=Path, required=True, help="Ruta del archivo XLSX con las reglas.")
    parser.add_argument(
        "--target",
        type=Path,
        default=Path("resource/quality_rules"),
        help="Carpeta destino para los archivos JSON generados.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    grouped = gather_rules(args.source)
    if not grouped:
        raise SystemExit("No se pudo extraer ninguna regla desde el archivo indicado.")
    written = write_outputs(args.target, grouped)
    for path in written:
        print(f"Escrib? {path}")


if __name__ == "__main__":
    main()
